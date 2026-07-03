import os
import platform
import re
import json
import openpyxl

# Repo root, derived from this file's location (src/tools/asset_index_builder.py)
# so every developer's own checkout resolves correctly, regardless of OS or path.
ROOT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "index.json")
TTS_INDEX_XLSX = os.path.join(os.path.dirname(__file__), "assets_tts_url.xlsx")
DRIVE_ROOT = "/Users/yurii/My Drive (pe.ur.ur@gmail.com)/Arkham Horror (TTS)"
ASSETS_ROOT = os.path.join(ROOT_DIR, "assets")

# Every google.drivePath category that has been migrated out of the Drive
# mount into the project's assets/ folder (see asset_migrator.py) — these
# take priority over any Drive copy when resolving google.local.
DRIVE_PATH_MAP = {
    "Arkham Horror (TTS)/(0) Base": "cards/0_base",
    "Arkham Horror (TTS)/(0) Base/EN": "cards/0_base/en",
    "Arkham Horror (TTS)/(0) Base/Ukr": "cards/0_base/uk",
    "Arkham Horror (TTS)/(0) Base/investigators_mini": "investigators/0_base/mini",
    "Arkham Horror (TTS)/(1) Dead of Night": "cards/1_dead_of_night",
    "Arkham Horror (TTS)/(1) Dead of Night/en": "cards/1_dead_of_night/en",
    "Arkham Horror (TTS)/(1) Dead of Night/Ukr": "cards/1_dead_of_night/uk",
    "Arkham Horror (TTS)/(1) Dead of Night/investigators_mini": "investigators/1_dead_of_night/mini",
    "Arkham Horror (TTS)/(2) Under Dark Waves": "cards/2_under_dark_waves",
    "Arkham Horror (TTS)/(2) Under Dark Waves/en": "cards/2_under_dark_waves/en",
    "Arkham Horror (TTS)/(2) Under Dark Waves/ukr": "cards/2_under_dark_waves/uk",
    "Arkham Horror (TTS)/(2) Under Dark Waves/investigators_mini": "investigators/2_under_dark_waves/mini",
    "Arkham Horror (TTS)/(3) Secrets of the Order": "cards/3_secrets_of_the_order",
    "Arkham Horror (TTS)/(3) Secrets of the Order/en": "cards/3_secrets_of_the_order/en",
    "Arkham Horror (TTS)/(3) Secrets of the Order/ukr": "cards/3_secrets_of_the_order/uk",
    "Arkham Horror (TTS)/(3) Secrets of the Order/investigators_mini": "investigators/3_secrets_of_the_order/mini",
    "Arkham Horror (TTS)/(4) Recursive Echoes/en": "cards/4_recursive_echoes/en",
    "Arkham Horror (TTS)/(4) Recursive Echoes/ukr": "cards/4_recursive_echoes/uk",
    "Arkham Horror (TTS)/Icons": "icons",
    "Arkham Horror (TTS)/Music": "music",
    "Arkham Horror (TTS)/Other graphics": "other_graphics",
    "Arkham Horror (TTS)/Other graphics/additional_buttons for xml": "other_graphics/additional_buttons_for_xml",
    "Arkham Horror (TTS)/investigator/(0) base": "investigators/0_base",
    "Arkham Horror (TTS)/investigator/(1) Dead of Night": "investigators/1_dead_of_night",
    "Arkham Horror (TTS)/investigator/(2) Under Dark Waves": "investigators/2_under_dark_waves",
    "Arkham Horror (TTS)/investigator/(3) Secrets of the Order": "investigators/3_secrets_of_the_order",
    "Arkham Horror (TTS)/map tiles": "map_tiles",
    "Arkham Horror (TTS)/map tiles/EN": "map_tiles/en",
    "Arkham Horror (TTS)/map tiles/Ukr": "map_tiles/uk",
    "Arkham Horror (TTS)/models": "models",
    "Arkham Horror (TTS)/rules/EN": "rules/en",
    "Arkham Horror (TTS)/rules/UKR": "rules/uk",
    "Arkham Horror (TTS)/ukrfonts": "fonts",
}

# One-time cleanup applied by asset_migrator.py --cleanup-cache: drops the
# library/ + cache/ wrapper folders entirely. library/* moves up to assets/*
# directly; the individual steam_cache/external_cache files (previously named
# by opaque Steam/hash id) get real names based on what actually references
# them, and land in the matching content folder instead of a "cache" bucket.
ASSET_FLATTEN_MOVES = (
    ("library/cards", "cards"),
    ("library/investigators", "investigators"),
    ("library/icons", "icons"),
    ("library/music", "music"),
    ("library/other_graphics", "other_graphics"),
    ("library/map_tiles", "map_tiles"),
    ("library/models", "models"),
    ("library/rules", "rules"),
    ("library/fonts", "fonts"),
)

CACHE_FILE_RENAMES = {
    # steam_cache
    "1004810215902243100.png": "other_graphics/sky_background.png",
    "1022822306775849601.obj": "models/clue.obj",
    "10793773223729217.png": "other_graphics/custom_tile_banner.png",
    "12672819282907927825.mp3": "music/cthulhu_music.mp3",
    "1616187808495853030.png": "models/blank_diffuse.png",
    "1626352406049452193.unity3d": "investigators/models/agatha_crane.unity3d",
    "1626353262602904348.unity3d": "investigators/models/zoey_samaras.unity3d",
    "1626353262602913272.unity3d": "investigators/models/minh_thi_phan.unity3d",
    "1628605470660039560.unity3d": "investigators/models/tommy_muldoon.unity3d",
    "1628605470660050553.unity3d": "investigators/models/mark_harrigan.unity3d",
    "1691654041388028254.unity3d": "investigators/models/charlie_kane.unity3d",
    "1693904066486102818.unity3d": "investigators/models/father_mateo.unity3d",
    "1693904088612102709.unity3d": "investigators/models/calvin_wright.unity3d",
    "1695028324684325544.unity3d": "investigators/models/preston_fairmont.unity3d",
    "1698402494141013039.unity3d": "investigators/models/kate_winthrop.unity3d",
    "1698402701139074953.unity3d": "investigators/models/rex_murphy.unity3d",
    "1698403597538910237.unity3d": "investigators/models/norman_withers.unity3d",
    "1698403941769899958.unity3d": "investigators/models/marie_lambeau.unity3d",
    "1698404049981089036.unity3d": "investigators/models/skids_otoole.unity3d",
    "1699531019170208342.unity3d": "investigators/models/agnes_baker.unity3d",
    "1699531490809791962.unity3d": "investigators/models/dexter_drake.unity3d",
    "1735566544368127960.obj": "models/map_tile_variant_a.obj",
    "1735566544368129028.obj": "models/map_tile_variant_a_collider.obj",
    "1735566544368132749.obj": "models/map_tile_variant_b.obj",
    "1735566544368133041.obj": "models/map_tile_variant_b_collider.obj",
    "1740071314870270469.unity3d": "investigators/models/carson_sinclair.unity3d",
    "1751308203937768602.obj": "models/infinite_bag.obj",
    "1781715244011610564.unity3d": "investigators/models/roland_banks.unity3d",
    "1807610360373912251.unity3d": "investigators/models/jenny_barnes.unity3d",
    "1809896678265162543.unity3d": "investigators/models/michael_mcglen.unity3d",
    "1809896946455457610.unity3d": "investigators/models/silas_marsh.unity3d",
    "1811019838027491786.unity3d": "investigators/models/patrice_hathaway.unity3d",
    "1811020692182300783.unity3d": "investigators/models/wendy_adams.unity3d",
    "1811021436411387289.unity3d": "investigators/models/daniela_reyes.unity3d",
    "1814396202014691275.unity3d": "investigators/models/diana_stanley.unity3d",
    "1820021533259458573.obj": "investigators/models/base_standee.obj",
    "1825645452880556651.unity3d": "investigators/models/ashcan_pete.unity3d",
    "1829040563499276403.unity3d": "investigators/models/stella_clark.unity3d",
    "2042982245664635665.unity3d": "investigators/models/winifred_habbamock.unity3d",
    "2050872233147596247.obj": "models/map_tile_hidden_path.obj",
    "2050872233147596437.obj": "models/map_tile_hidden_path_collider.obj",
    "87098596225685535.obj": "models/mythos_cup.obj",
    "922542758751649800.obj": "models/model_base.obj",
    "959719855127875098.jpg": "models/mythos_cup_diffuse.jpg",
    # external_cache
    "2c439e30feb136ecdde637d998553021.pdf": "rules/secrets_of_the_order_rulebook.pdf",
    "6af8063db625f2d6f8e57131659ae386.pdf": "rules/quick_reference.pdf",
    "e9069282359a69ebffb0e5ef3953d1f0.pdf": "rules/base_rules_reference.pdf",
    "ebee3af0c2180e91006fff6051f3541b.pdf": "rules/rulebook.pdf",
}

URL_REGEX = re.compile(r'https?://[^\s"\']+')


# -------------------------
# LOCAL FILE URL / PATH HELPERS (cross-platform)
# -------------------------
def build_local_file_url(abs_path: str) -> str:
    if platform.system() == "Windows":
        return "file:///" + abs_path.replace("\\", "/")
    return "file:////" + abs_path.lstrip("/")


def to_relative(abs_path):
    if not abs_path:
        return abs_path

    try:
        rel = os.path.relpath(abs_path, ROOT_DIR)
    except ValueError:
        return abs_path

    if rel.startswith(".."):
        return abs_path

    return rel.replace(os.sep, "/")


def to_absolute(rel_path: str) -> str:
    return os.path.join(ROOT_DIR, *rel_path.split("/"))


def path_from_file_url(url: str):
    if not url.lower().startswith("file:"):
        return None

    remainder = re.sub(r"^file:/+", "", url, flags=re.IGNORECASE)

    if re.match(r"^[a-zA-Z]:", remainder):
        return remainder.replace("/", "\\")

    return "/" + remainder


def assets_subpath(url_or_path: str):
    """Given a file:// URL or raw path, return whatever comes after the last
    'assets' path segment (any OS separator), for re-rooting onto this
    machine's own checkout. Returns None if there's no 'assets' segment."""
    path = path_from_file_url(url_or_path) or url_or_path
    normalized = path.replace("\\", "/")

    marker = "/assets/"
    idx = normalized.rfind(marker)
    if idx == -1:
        return None

    return normalized[idx + len(marker):]


# -------------------------
# NORMALIZE
# -------------------------
def normalize(text: str) -> str:
    return (
        text.replace("\\u0026", "&")
            .replace("\\/", "/")
    )


# -------------------------
# FIX URLS
# -------------------------
def fix_url(url: str) -> str:
    return (
        url.replace("httpssteam", "https://steam")
           .replace("httpsteam", "http://steam")
           .replace("httpscloud", "https://cloud")
           .replace("httpcloud", "http://cloud")
           .replace("httpssteamusercontent", "https://steamusercontent")
           .replace("httpsteamusercontent", "http://steamusercontent")
    )


# -------------------------
# SOURCE DETECTION
# -------------------------
def classify(url: str):
    if "drive.google.com" in url:
        return "google"
    if "steamusercontent" in url:
        return "steam"
    return "other"


# -------------------------
# ID EXTRACTORS
# -------------------------
def extract_google_id(url: str):
    import re
    m = re.search(r"/d/([a-zA-Z0-9_-]+)", url)
    if m:
        return m.group(1)

    m = re.search(r"[?&]id=([a-zA-Z0-9_-]+)", url)
    if m:
        return m.group(1)

    return None


def extract_steam_id(url: str):
    import re
    m = re.search(r"ugc/([0-9]+)", url)
    if m:
        return m.group(1)

    nums = re.findall(r"([0-9]{8,})", url)
    return nums[-1] if nums else None


# -------------------------
# LOAD TTS_INDEX SHEET (id -> name, drivePath)
# -------------------------
def load_tts_index():
    if not os.path.exists(TTS_INDEX_XLSX):
        return {}

    wb = openpyxl.load_workbook(TTS_INDEX_XLSX, data_only=True)
    if "TTS_Index" not in wb.sheetnames:
        return {}

    ws = wb["TTS_Index"]

    lookup = {}
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # skip header

    for row_id, name, drive_path, url in rows:
        if row_id:
            lookup[row_id] = {"name": name, "drivePath": drive_path}

    return lookup


# -------------------------
# LOAD LOCAL DRIVE FILES (drivePath -> {stem: full_path})
# -------------------------
def load_local_files():
    local_files = {}

    if os.path.isdir(DRIVE_ROOT):
        drive_parent = os.path.dirname(DRIVE_ROOT)

        for dirpath, _, files in os.walk(DRIVE_ROOT):
            drive_path = os.path.relpath(dirpath, drive_parent)
            stems = {}

            for f in files:
                full_path = os.path.join(dirpath, f)
                stem, _ = os.path.splitext(f)
                stems[stem] = full_path
                stems[f] = full_path

            local_files[drive_path] = stems

    # Everything migrated into assets/ (see asset_migrator.py) also takes
    # priority over its old Drive location.
    for drive_path, mapped in DRIVE_PATH_MAP.items():
        asset_dir = os.path.join(ASSETS_ROOT, mapped)
        if not os.path.isdir(asset_dir):
            continue

        stems = {}
        for f in os.listdir(asset_dir):
            full_path = os.path.join(asset_dir, f)
            if not os.path.isfile(full_path):
                continue
            stem, _ = os.path.splitext(f)
            stems[stem] = full_path
            stems[f] = full_path

        local_files[drive_path] = stems

    return local_files


def fill_local_path(asset_google, local_files):
    drive_path = asset_google.get("drivePath")
    name = asset_google.get("name")

    if not drive_path or not name:
        return

    asset_google["local"] = local_files.get(drive_path, {}).get(name)


# -------------------------
# LOAD "Files" SHEET REVERSE (google id -> project-relative local path)
# TTS_Index (name/drivePath based resolution) no longer exists in the xlsx;
# this is the fallback for resolving brand-new raw google links that were
# never localized, using the same "Files" sheet add_download_links.py reads.
# -------------------------
def load_files_index():
    if not os.path.exists(TTS_INDEX_XLSX):
        return {}

    wb = openpyxl.load_workbook(TTS_INDEX_XLSX, data_only=True)
    if "Files" not in wb.sheetnames:
        return {}

    ws = wb["Files"]
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # skip header

    lookup = {}
    for path, link in rows:
        if not path or not link:
            continue
        gid = extract_google_id(link)
        if gid:
            lookup[gid] = path

    return lookup


def fill_local_path_from_files_index(asset_google, files_index):
    if asset_google.get("local"):
        return

    rel_path = files_index.get(asset_google.get("id"))
    if rel_path:
        asset_google["local"] = to_absolute(rel_path)


# -------------------------
# LOAD EXISTING INDEX (MERGE SUPPORT)
# -------------------------
def load_index():
    if not os.path.exists(OUTPUT_FILE):
        return {}

    try:
        with open(OUTPUT_FILE, "r", encoding="utf-8") as f:
            return json.load(f).get("data", {})
    except:
        return {}


# -------------------------
# SCHEMA FIX / INFLATE (load-time: restore absolute paths + the internal
# google/steam/other sub-dicts so add_asset()'s merge-by-target logic always
# works with a uniform shape, regardless of whether the on-disk index.json is
# the old nested format or the new flat one written by prune_for_output())
# -------------------------
def ensure_schema(index):
    inflated = {}

    for file_path, file_data in index.items():
        abs_path = file_path if os.path.isabs(file_path) else to_absolute(file_path)
        assets = file_data.get("assets", [])

        for asset in assets:
            local = asset.pop("local", None)
            # _gurl is carried forward verbatim as a re-download reference for
            # an already-local file — deliberately kept separate from the
            # google/steam sub-dicts below, which specifically mean "this
            # target's live text still has a raw drive.google.com URL to
            # localize." Reusing google.link/local for gUrl would make
            # already-migrated assets look like they still need migrating.
            asset["_gurl"] = asset.pop("gUrl", None)

            abs_local = None
            if local:
                abs_local = local if os.path.isabs(local) else to_absolute(local)

            google = asset.setdefault("google", {})
            for key in ("id", "name", "link", "local", "drivePath"):
                google.setdefault(key, None)
            if google.get("local") and not os.path.isabs(google["local"]):
                google["local"] = to_absolute(google["local"])

            steam = asset.setdefault("steam", {})
            for key in ("id", "name", "link", "local"):
                steam.setdefault(key, None)
            if steam.get("local") and not os.path.isabs(steam["local"]):
                steam["local"] = to_absolute(steam["local"])

            other = asset.setdefault("other", {})
            other.setdefault("link", None)

            if abs_local:
                other["link"] = build_local_file_url(abs_local)

        inflated[abs_path] = {"path": abs_path, "assets": assets}

    return inflated


# -------------------------
# PRUNE FOR OUTPUT (write-time): flatten to {target, local, gUrl} — drop the
# internal google/steam/other split, drop null fields, and turn absolute
# paths under ROOT_DIR into relative ones so index.json is portable and easy
# to read: just "where's the file locally" + "where to re-download it from".
# -------------------------
def prune_for_output(index):
    output = {}

    for file_path, file_data in index.items():
        rel_file_path = to_relative(file_path)
        assets_out = []

        for asset in file_data.get("assets", []):
            google = asset.get("google", {})
            steam = asset.get("steam", {})
            other_link = asset.get("other", {}).get("link")

            local = None
            gurl = asset.get("_gurl") or google.get("link")

            if google.get("local"):
                local = to_relative(google["local"])
            elif steam.get("local"):
                local = to_relative(steam["local"])
            elif other_link and other_link.lower().startswith("file:"):
                abs_path = path_from_file_url(other_link)
                local = to_relative(abs_path) if abs_path else None

            asset_out = {"target": asset["target"]}
            if local:
                asset_out["local"] = local
            if gurl:
                asset_out["gUrl"] = gurl

            if len(asset_out) > 1:
                assets_out.append(asset_out)

        output[rel_file_path] = {"path": rel_file_path, "assets": assets_out}

    return output


# -------------------------
# ADD / MERGE ASSET
# -------------------------
def fill_google_metadata(asset_google, tts_index):
    info = tts_index.get(asset_google["id"])
    if info:
        asset_google["name"] = info["name"]
        asset_google["drivePath"] = info["drivePath"]


def add_asset(index, file_path, source, url, target, tts_index, local_files, files_index):
    if file_path not in index:
        index[file_path] = {
            "path": file_path,
            "assets": []
        }

    assets = index[file_path]["assets"]

    # find existing by target
    for asset in assets:
        if asset["target"] == target:
            asset[source]["link"] = url

            if source == "google":
                asset[source].setdefault("drivePath", None)
                asset[source]["id"] = extract_google_id(url)
                fill_google_metadata(asset[source], tts_index)
                fill_local_path(asset[source], local_files)
                fill_local_path_from_files_index(asset[source], files_index)

            if source == "steam":
                asset[source]["id"] = extract_steam_id(url)

            return

    # create new asset
    new_asset = {
        "target": target,

        "google": {
            "id": None,
            "name": None,
            "link": None,
            "local": None,
            "drivePath": None
        },

        "steam": {
            "id": None,
            "name": None,
            "link": None,
            "local": None
        },

        "other": {
            "link": None
        }
    }

    new_asset[source]["link"] = url

    if source == "google":
        new_asset[source]["id"] = extract_google_id(url)
        fill_google_metadata(new_asset[source], tts_index)
        fill_local_path(new_asset[source], local_files)
        fill_local_path_from_files_index(new_asset[source], files_index)

    elif source == "steam":
        new_asset[source]["id"] = extract_steam_id(url)

    assets.append(new_asset)


# -------------------------
# JSON WALKER
# -------------------------
def walk(obj, file_path, index, tts_index, local_files, files_index, path=None):
    if path is None:
        path = []

    if isinstance(obj, dict):
        for k, v in obj.items():
            walk(v, file_path, index, tts_index, local_files, files_index, path + [k])

    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            walk(v, file_path, index, tts_index, local_files, files_index, path + [i])

    elif isinstance(obj, str):
        if "http" not in obj and "file:" not in obj:
            return

        url = fix_url(normalize(obj))
        source = classify(url)

        add_asset(
            index,
            file_path,
            source,
            url,
            path,
            tts_index,
            local_files,
            files_index
        )


# -------------------------
# SCAN PROJECT
# -------------------------
def scan():
    index = ensure_schema(load_index())
    tts_index = load_tts_index()
    local_files = load_local_files()
    files_index = load_files_index()

    IGNORE_DIRS = {
        os.path.realpath(os.path.join(ROOT_DIR, ".tts")),
        os.path.realpath(os.path.join(ROOT_DIR, "src/tools")),
    }

    def ignored(path):
        path = os.path.realpath(path)
        return any(path.startswith(i) for i in IGNORE_DIRS)

    for root, dirs, files in os.walk(ROOT_DIR):

        dirs[:] = [
            d for d in dirs
            if not ignored(os.path.join(root, d))
        ]

        if ignored(root):
            continue

        for file in files:
            if not file.endswith((".json", ".lua", ".xml", ".txt")):
                continue

            file_path = os.path.join(root, file)

            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    content = json.load(f)
            except:
                continue

            walk(content, file_path, index, tts_index, local_files, files_index)

    # Drop stale entries for files that were renamed/deleted since the last
    # indexed run — otherwise merge-on-load_index() keeps them forever.
    for stale_path in [p for p in index if not os.path.isfile(p)]:
        del index[stale_path]

    return index


# -------------------------
# MAIN
# -------------------------
def main():
    index = scan()

    output = {}
    if os.path.exists(OUTPUT_FILE):
        try:
            with open(OUTPUT_FILE, "r", encoding="utf-8") as f:
                output = json.load(f)
        except (json.JSONDecodeError, OSError):
            output = {}

    output["data"] = prune_for_output(index)

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print("DONE")
    print("Files:", len(index))


if __name__ == "__main__":
    main()